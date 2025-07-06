from fastapi import UploadFile, Form
import openpyxl
from io import BytesIO
from datetime import datetime
import httpx
import openpyxl.utils
from fastapi import Header

TIPOS_PERMITIDOS = {
    "Auxilio de transporte": "SOLICITUDES.xlsx",
    "Descuento": "SOLICITUDES.xlsx",
    "Otros": "SOLICITUDES.xlsx",
    "Otro Si Definitivo": "SOLICITUDES.xlsx",
    "Horas Extra": "SOLICITUDES2.xlsx",
    "Otro Si Temporal": "SOLICITUDES3.xlsx",
    "Vacaciones": "SOLICITUDES4.xlsx",
}

def validar_horas_extra(fila, row_idx, campos_obligatorios, errores):
    fila_valida = True

    concepto_idx = campos_obligatorios.get("CONCEPTO")
    codigo_idx = campos_obligatorios.get("CON_CODIGO")
    unidad_idx = campos_obligatorios.get("UNIDADES")

    # ✅ Extraer valores
    concepto = str(fila[concepto_idx].value).strip() if concepto_idx is not None else ""
    codigo = str(fila[codigo_idx].value).strip() if codigo_idx is not None else ""
    unidad = str(fila[unidad_idx].value).strip() if unidad_idx is not None else ""

    concepto_valido_map = {
        "Domingo Sin Compensatorio Diurno": "75",
        "Domingo Sin Compensatorio Nocturno": "110",
        "Dominical Con Compensatorio Diurno": "66",
        "Dominical Con Compensatorio Nocturno": "78",
        "Festivo Sin Compensatorio Diurno": "75",
        "Hora extra Diurna": "55",
        "Recargo Nocturno 35%": "45",
    }

    # ✅ Validar CONCEPTO
    if concepto not in concepto_valido_map:
        errores.append(
            f"❌ Fila {row_idx}, Columna {concepto_idx + 1} (CONCEPTO): \"{concepto}\" no es válido. Verifica que esté escrito correctamente según las opciones disponibles."
        )
        fila_valida = False
    else:
        # ✅ Validar que el código corresponde al concepto
        codigo_esperado = concepto_valido_map[concepto]
        if codigo != codigo_esperado:
            errores.append(
                f"❌ Fila {row_idx}, Columna {codigo_idx + 1} (CON_CODIGO): se esperaba \"{codigo_esperado}\" para el concepto \"{concepto}\", pero llegó \"{codigo}\"."
            )
            fila_valida = False

    # ✅ Validar UNIDADES
    try:
        float(unidad)
    except Exception:
        errores.append(
            f"❌ Fila {row_idx}, Columna {unidad_idx + 1} (UNIDADES): debe ser un número válido (puede tener decimales). Valor ingresado: \"{unidad}\"."
        )
        fila_valida = False

    return fila_valida


VALIDACIONES_ESPECIALES = {
    "Horas Extra": validar_horas_extra,
}

def normalizar_fecha(fecha_raw):
    if isinstance(fecha_raw, datetime):
        # Forzar hora a las 05:00:00
        return fecha_raw.replace(hour=5, minute=0, second=0, microsecond=0)
    elif isinstance(fecha_raw, str):
        try:
            fecha = datetime.strptime(fecha_raw.strip(), "%d/%m/%Y")
            return fecha.replace(hour=5, minute=0, second=0, microsecond=0)
        except Exception:
            return None
    return None

async def validar_duplicado_en_backend(cedula: str, fecha: datetime, tipo: str, nombre: str, jwt_token: str):
    url = "http://localhost:3000/novedad/validar-duplicado"

    fecha_str = fecha.strftime("%Y-%m-%d %H:%M:%S")
    params = {
        "cedula": cedula,
        "fecha": fecha_str,
        "tipo": tipo,
        "nombre": nombre
    }

    # 👇 Asegura que el token tenga el prefijo "Bearer "
    if jwt_token and not jwt_token.lower().startswith("bearer "):
        jwt_token = f"Bearer {jwt_token}"

    headers = {
        "Authorization": jwt_token
    }

    print(f"🔍 [MICROSERVICIO] Enviando request a BD: {params}")

    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url, params=params, headers=headers, timeout=10)
            print(f"🔍 [MICROSERVICIO] Status code: {response.status_code}")
            print(f"🔍 [MICROSERVICIO] Response text: {response.text}")
            
            if response.status_code == 200:
                data = response.json()
                print(f"🔍 [MICROSERVICIO] Response data: {data}")
                existe = data.get("existe", False)
                mensaje = data.get("mensaje", "")
                print(f"🔍 [MICROSERVICIO] ¿Existe duplicado? {existe}")
                return existe, mensaje
            else:
                print(f"❌ [MICROSERVICIO] Error HTTP: {response.status_code}")
                return False, f"⚠️ Error al consultar duplicado: código {response.status_code}"
    except Exception as e:
        print(f"❌ [MICROSERVICIO] Excepción: {e}")
        return False, f"⚠️ Excepción al validar duplicado: {e}"


async def validar_excel(
    file: UploadFile,
    tipo: str,
    titulo: str = Form(...),
    nombreUsuario: str = Form(...),
    nombreTienda: str = Form(...),
    authorization: str = Header(None), 
):
    
    print(f"📥 Archivo recibido: {file.filename}")
    print(f"📋 Tipo: {tipo}, Título: {titulo}, Usuario: {nombreUsuario}, Tienda: {nombreTienda}")
    
    errores = []
    cantidad_solicitudes = 0  # Contador de filas válidas

    if tipo not in TIPOS_PERMITIDOS:
        return {
            "valido": False,
            "errores": [f"❌ Tipo de solicitud '{tipo}' no es válido."]
        }

    content = await file.read()
    wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    sheet = wb.active

    encabezados = [cell.value for cell in sheet[5]]
    encabezados_normalizados = [str(h).strip().upper() if h else "" for h in encabezados]
    print("📌 Encabezados detectados:", encabezados_normalizados)

    # Diccionario de cabeceras esperadas por tipo
    cabeceras_por_tipo = {
        "SOLICITUDES.xlsx": [
            "N", "FECHA DE REPORTE", "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)",
            "CATEGORIA", "TIENDA", "QUIEN REPORTA LA NOVEDAD\n(Nombre Jefe GH)", "DETALLE NOVEDAD"
        ],
        "SOLICITUDES2.xlsx": [
            "N", "FECHA DE REPORTE", "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)",
            "CATEGORIA", "TIENDA", "QUIEN REPORTA LA NOVEDAD\n(Nombre Jefe GH)", "DETALLE NOVEDAD",
            "CONCEPTO", "CON_CODIGO", "UNIDADES", "FECHA NOVEDAD"
        ],
        "SOLICITUDES3.xlsx": [
            "N", "FECHA DE REPORTE", "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)",
            "CATEGORIA", "TIENDA", "QUIEN REPORTA LA NOVEDAD\n(Nombre Jefe GH)", "DETALLE NOVEDAD",
            "JORNADA EMPLEADO", "JORNADA OTRO SI TEMPORAL", "FECHA INICIO", "FECHA FIN",
            "SALARIO ACTUAL", "SALARIO OTRO SI TEMPORAL", "CONSECUTIVO FORMS"
        ],
        "SOLICITUDES4.xlsx": [
            "N", "FECHA DE REPORTE", "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)",
            "CATEGORIA", "TIENDA", "QUIEN REPORTA LA NOVEDAD\n(Nombre Jefe GH)", "DETALLE NOVEDAD",
            "DIAS A TOMAR", "FECHA INICIO", "FECHA FIN"
        ]
    }
    
    CAMPOS_OBLIGATORIOS_POR_PLANTILLA = {
        "SOLICITUDES.xlsx": [
            "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)", "DETALLE NOVEDAD"
        ],
        "SOLICITUDES2.xlsx": [
            "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)", "DETALLE NOVEDAD",
            "CONCEPTO", "CON_CODIGO", "UNIDADES", "FECHA NOVEDAD"
        ],
        "SOLICITUDES3.xlsx": [
            "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)", "DETALLE NOVEDAD",
            "JORNADA EMPLEADO", "JORNADA OTRO SI TEMPORAL", "FECHA INICIO", "FECHA FIN",
            "SALARIO ACTUAL", "SALARIO OTRO SI TEMPORAL", "CONSECUTIVO FORMS"
        ],
        "SOLICITUDES4.xlsx": [
            "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)", "DETALLE NOVEDAD",
            "DIAS A TOMAR", "FECHA INICIO", "FECHA FIN"
        ]
    }

    plantilla_esperada = TIPOS_PERMITIDOS[tipo]
    cabeceras_esperadas = [c.strip().upper() for c in cabeceras_por_tipo[plantilla_esperada]]
    
    #SOLO COLUMNAS NECESARIAS
    encabezados_truncados = encabezados_normalizados[:len(cabeceras_esperadas)]

    # Comparar con los headers reales del archivo
    if encabezados_truncados != cabeceras_esperadas:
        errores.append("❌ Las cabeceras no coinciden con el formato esperado.")
        errores.append(f"🔎 Esperado: {cabeceras_esperadas}")
        errores.append(f"📄 Recibido: {encabezados_truncados}")

    if errores:
        print("🛑 Errores de encabezado:", errores)
        return {
            "valido": False,
            "errores": errores
        }
        
    #VALIDACIONES EXTRA
    
    #1.Datos fuera del rango
    max_col_permitida = len(cabeceras_esperadas)
    
    for row_idx in range(6, sheet.max_row +1):
        fila = sheet[row_idx]
        for col_idx in range(max_col_permitida, sheet.max_column): #Columnas fuera de rango
            if col_idx >= len(fila):
                continue
            valor_extra = fila[col_idx].value
            if valor_extra is not None and str(valor_extra).strip() != "":
                letra_col = openpyxl.utils.get_column_letter(col_idx + 1)
                errores.append(
                     f"❌ Fila {row_idx}, columna {letra_col}: No debe contener información. Solo se permiten columnas hasta la {openpyxl.utils.get_column_letter(max_col_permitida)}."
                )

    # Construir los campos obligatorios para verificación de contenido 
    campos_obligatorios = {}
    
    campos_esperados = CAMPOS_OBLIGATORIOS_POR_PLANTILLA[plantilla_esperada]
    
    for campo in campos_esperados:
        if campo in encabezados_normalizados:
            campos_obligatorios[campo] = encabezados_normalizados.index(campo)
        else:
            errores.append(f"❌ Falta la columna obligatoria: {campo}")

    if errores:
        print("🛑 Errores por campos obligatorios:", errores)
        return {
            "valido": False,
            "errores": errores
        }

    # SET CONTROL DE DUPLICADOS EN LA MISMA PLANTILLA
    duplicados_cedula_fecha = set()
    
    # MEJORADO: Validar todos los registros contra la BD al inicio
    registros_para_validar = []
    
    # Primero recopilar todos los registros válidos
    for row_idx in range(6, sheet.max_row + 1):
        fila = sheet[row_idx]
        fila_visible = any(cell.value is not None for cell in fila)

        if not fila_visible:
            continue

        # Extraer datos principales
        cedula_idx = campos_obligatorios.get("CEDULA")
        fecha_idx = 1  # COLUMNA B --> FECHA
        
        cedula_val = str(fila[cedula_idx].value).strip() if cedula_idx is not None else ""
        fecha_val = fila[fecha_idx].value
        
        if isinstance(fecha_val, datetime):
            fecha_val = fecha_val.replace(hour=5, minute=0, second=0, microsecond=0)
        elif isinstance(fecha_val, str):
            try:
                fecha_val = datetime.strptime(fecha_val.strip(), "%d/%m/%Y").replace(hour=5, minute=0, second=0, microsecond=0)
            except Exception:
                fecha_val = None
        
        if cedula_val and fecha_val:
            registros_para_validar.append({
                'row_idx': row_idx,
                'cedula': cedula_val,
                'fecha': fecha_val,
                'tipo': tipo
            })

    # Validar duplicados en BD de forma masiva
    print(f"🔍 [MICROSERVICIO] Validando {len(registros_para_validar)} registros contra la BD...")
    duplicados_bd = set()
    
    for registro in registros_para_validar:
        print(f"🔍 [MICROSERVICIO] Validando registro: {registro}")
        
        jwt_token = str(authorization)
        
        existe_en_bd, mensaje_duplicado = await validar_duplicado_en_backend(
            registro['cedula'],
            registro['fecha'],
            registro['tipo'],
            "",  # nombre no es necesario para la validación de duplicados
            jwt_token 
        )
        
        print(f"🔍 [MICROSERVICIO] Resultado para {registro['cedula']}: existe={existe_en_bd}, mensaje={mensaje_duplicado}")
        
        if existe_en_bd:
            # Crear clave única para identificar el duplicado
            clave_duplicado = f"{registro['cedula']}-{registro['fecha'].strftime('%Y-%m-%d')}-{registro['tipo']}"
            duplicados_bd.add(clave_duplicado)
            print(f"⚠️ [MICROSERVICIO] Duplicado encontrado en BD: {clave_duplicado}")

    print(f"🔍 [MICROSERVICIO] Total duplicados encontrados en BD: {len(duplicados_bd)}")
    
    # Ahora procesar todas las filas
    for row_idx in range(6, sheet.max_row + 1):
        fila = sheet[row_idx]
        fila_visible = any(cell.value is not None for cell in fila)

        if not fila_visible:
            print(f"⚪ Fila {row_idx} vacía. Saltando...")
            continue

        print(f"\n🔍 Validando fila {row_idx}...")
        fila_valida = True

        # Validar campos obligatorios
        for campo, col_idx in campos_obligatorios.items():
            cell = fila[col_idx]
            valor = cell.value

            if valor is None or str(valor).strip() == "":
                mensaje_error = f"❌ Fila {row_idx}, Columna {col_idx + 1} ({campo}): VACÍA"
                print(mensaje_error)
                errores.append(mensaje_error)
                fila_valida = False
            else:
                print(f"✅ Fila {row_idx}, Columna {col_idx + 1} ({campo}): OK → '{valor}'")
            
                # VALIDACIONES EXTRA
                
                # 3. LONGITUD DE MENSAJE EN DETALLE NOVEDAD
                if campo == "DETALLE NOVEDAD":
                    texto = str(valor).strip()
                    longitud = len(texto)
                    if longitud < 15 or longitud > 250:
                        errores.append(
                            f"❌ Fila {row_idx}, Columna {col_idx + 1} (DETALLE NOVEDAD): debe tener entre 15 y 250 caracteres. Tiene {longitud}."
                        )
                        fila_valida = False
                
                # 4. SIN CARACTERES RAROS EN NOMBRE
                if campo == "NOMBRE (APELLIDOS-NOMBRES)":
                    texto = str(valor).strip()
                    if not all(char.isalpha() or char.isspace() or char in "áéíóúÁÉÍÓÚñÑ" for char in texto):
                        errores.append(
                            f"❌ Fila {row_idx}, Columna {col_idx + 1} (NOMBRE): contiene caracteres inválidos. Solo se permiten letras, tildes y espacios."
                        )
                        fila_valida = False
                
                # 5. CAMPO CEDULA SOLO NUMEROS
                if campo == "CEDULA":
                    texto = str(valor).strip()
                    if not texto.isdigit():
                        errores.append(
                            f"❌ Fila {row_idx}, Columna {col_idx + 1} (CÉDULA): debe contener solo números. Valor recibido: '{texto}'."
                        )
                        fila_valida = False

        # Validaciones automáticas por columnas generadas
        cell_A = fila[0]  # Columna A
        cell_B = fila[1]  # Columna B
        cell_E = fila[4]  # Columna E
        cell_F = fila[5]  # Columna F
        cell_G = fila[6]  # Columna G

        # A: Número secuencial
        numero_esperado = row_idx - 5
        if cell_A.value != numero_esperado:
            errores.append(f"❌ Fila {row_idx}, columna A: Se esperaba el número '{numero_esperado}' y llegó '{cell_A.value}'")
            fila_valida = False

        # B: Fecha del día
        fecha_b = normalizar_fecha(cell_B.value)
        hoy = normalizar_fecha(datetime.now())
        if fecha_b != hoy:
            errores.append(f"❌ Fila {row_idx}, columna B: Se esperaba la fecha '{hoy.strftime('%d/%m/%Y')}' y llegó '{cell_B.value}'")
            fila_valida = False

        # E: Título
        if str(cell_E.value).strip() != titulo.strip():
            errores.append(f"❌ Fila {row_idx}, columna E: Se esperaba el título '{titulo}' y llegó '{cell_E.value}'")
            fila_valida = False

        # F: Tienda
        if str(cell_F.value).strip() != nombreTienda.strip():
            errores.append(f"❌ Fila {row_idx}, columna F: Se esperaba la tienda '{nombreTienda}' y llegó '{cell_F.value}'")
            fila_valida = False

        # G: Jefe
        if str(cell_G.value).strip() != nombreUsuario.strip():
            errores.append(f"❌ Fila {row_idx}, columna G: Se esperaba el nombre del jefe '{nombreUsuario}' y llegó '{cell_G.value}'")
            fila_valida = False
        
        # VALIDACIONES EXTRA
        
        # 2. No puede estar la misma persona el mismo día en la misma solicitud (dentro del archivo)
        cedula_idx = campos_obligatorios.get("CEDULA")
        fecha_idx = 1  # COLUMNA B --> FECHA
        
        cedula_val = str(fila[cedula_idx].value).strip() if cedula_idx is not None else ""
        fecha_val = fila[fecha_idx].value
        
        if isinstance(fecha_val, datetime):
            fecha_val = fecha_val.replace(hour=5, minute=0, second=0, microsecond=0)
        elif isinstance(fecha_val, str):
            try:
                fecha_val = datetime.strptime(fecha_val.strip(), "%d/%m/%Y").replace(hour=5, minute=0, second=0, microsecond=0)
            except Exception:
                fecha_val = None
        
        # Clave para duplicados en el mismo archivo
        clave_archivo = f"{cedula_val}-{fecha_val}"
        
        if clave_archivo in duplicados_cedula_fecha:
            errores.append(
                f"❌ Fila {row_idx}: La persona con cédula {cedula_val} ya tiene una solicitud registrada el día {fecha_val.strftime('%d/%m/%Y') if fecha_val else 'fecha inválida'} en este archivo."
            )
            fila_valida = False
        else:
            duplicados_cedula_fecha.add(clave_archivo)
        
        # VALIDAR DUPLICADOS EN BD (usando los datos pre-validados)
        if fecha_val:  # Solo validar si la fecha es válida
            clave_bd = f"{cedula_val}-{fecha_val.strftime('%Y-%m-%d')}-{tipo}"
            print(f"🔍 [MICROSERVICIO] Verificando clave BD: {clave_bd}")
            print(f"🔍 [MICROSERVICIO] Duplicados BD encontrados: {duplicados_bd}")
            
            if clave_bd in duplicados_bd:
                mensaje_error = f"❌ Fila {row_idx}: Ya existe una novedad en la base de datos con cédula {cedula_val}, fecha {fecha_val.strftime('%d/%m/%Y')} y tipo '{tipo}'"
                print(f"⚠️ [MICROSERVICIO] {mensaje_error}")
                errores.append(mensaje_error)
                fila_valida = False
            else:
                print(f"✅ [MICROSERVICIO] No hay duplicado en BD para: {clave_bd}")
        
        if tipo in VALIDACIONES_ESPECIALES and fila_valida:
            funcion_validadora = VALIDACIONES_ESPECIALES[tipo]
            resultado_especial = funcion_validadora(fila, row_idx, campos_obligatorios, errores)
            if not resultado_especial:
                print(f"⚠️ Fila {row_idx} falló la validación especial para tipo '{tipo}'")
                fila_valida = False
        
        if fila_valida:
            print(f"✅ Fila {row_idx} completa y válida.")
            cantidad_solicitudes += 1
        else:
            print(f"⚠️ Fila {row_idx} tiene errores.")

    if errores:
        print("\n🛑 Validación terminada con errores.")
        print(f"🛑 Total errores: {len(errores)}")
        return {
            "valido": False,
            "errores": errores,
            "tipoValidado": tipo,
            "cantidadSolicitudes": cantidad_solicitudes
        }

    print(f"\n✅ Validación exitosa. Total solicitudes válidas: {cantidad_solicitudes}")
    return {
        "valido": True,
        "esMasiva": True,
        "cantidadSolicitudes": cantidad_solicitudes,
        "tipoValidado": tipo
    }